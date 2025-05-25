# ---------------------------------------------------------
# Missile Booster Web App
# Authored by AChalli with AI-assisted coding review.
# All core design and final implementation by AChalli.
# ---------------------------------------------------------

import numpy as np
from flask import Flask, render_template, request, send_file, Response
from openpyxl import Workbook
import io
import threading
import webbrowser
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt

import base64
from io import BytesIO

import os, sys

if getattr(sys, 'frozen', False):
    # we’re running in a PyInstaller bundle
    base_path = sys._MEIPASS
else:
    # running in normal Python environment
    base_path = os.path.abspath(".")

template_dir = os.path.join(base_path, "templates")
app = Flask(__name__,
            template_folder=template_dir)


def calculate_rocket_parameters(payload_kg, diameter_m, length_m,
                                propellant_density, isp,
                                booster_lengths, propellant_fractions,
                                theta_deg, mode='standard', burn_time=0):
    # Constants
    gravity_mps = 9.81
    radius_m = diameter_m / 2
    num_boosters = len(booster_lengths)

    # 1) volumes & propellant masses
    volumes = [np.pi * (radius_m**2) * h for h in booster_lengths]
    propellant_masses = [V * propellant_density for V in volumes]

    # 2) dry (structure) masses
    structural_masses = [
        (mp / propellant_fractions[i]) - mp
        for i, mp in enumerate(propellant_masses)
    ]
    structural_masses[-1] += payload_kg  # last‐stage carries the payload

    # 3) build up m0 & mf for each stage
    initial_masses = []
    final_masses   = []
    current_mass   = sum(propellant_masses) + sum(structural_masses)

    for i in range(num_boosters):
        # before burning stage i
        initial_masses.append(current_mass + payload_kg if i==0 else current_mass)
        # after burning stage i
        final_masses.append(initial_masses[-1] - propellant_masses[i])
        # drop stage i’s dry+propellant
        current_mass -= (propellant_masses[i] + structural_masses[i])

    # 4) ΔV per stage (special‐case first stage if in burntime mode)
    delta_v_stages = []
    for i in range(num_boosters):
        m0 = initial_masses[i]
        mf = final_masses[i]
        if i == 0 and mode == 'burntime':
            # V = t_b·g + Isp·g·ln(m0/mf)
            v = burn_time * gravity_mps + isp * gravity_mps * np.log(m0/mf)
        else:
            # classic rocket eqn
            v = isp * gravity_mps * np.log(m0/mf)
        delta_v_stages.append(v)

    # 💥 Here’s the missing line you need:
    total_delta_v = sum(delta_v_stages)

    # 5) range & flight time
    theta_rad = np.radians(theta_deg)
    range_km   = (total_delta_v**2 / gravity_mps) * np.sin(2*theta_rad) / 1000
    flight_time = 2 * total_delta_v * np.sin(theta_rad) / gravity_mps

    return {
        'stage_delta_v':    delta_v_stages,
        'total_delta_v':    total_delta_v,
        'range_km':         range_km,
        'flight_time':      flight_time,
        'initial_masses':   initial_masses,
        'propellant_masses':propellant_masses,
        'structural_masses':structural_masses
    }



def generate_trajectory_graph(v_max, theta_deg):
    """Generate x-y trajectory graph"""
    g = 9.81
    theta_rad = np.radians(theta_deg)

    # Calculate flight time
    flight_time = (2 * v_max * np.sin(theta_rad)) / g

    # Create time points
    t = np.linspace(0, flight_time, 500)

    # Calculate x and y components of trajectory
    v_x = v_max * np.cos(theta_rad)
    x = v_x * t
    y = v_max * np.sin(theta_rad) * t - 0.5 * g * t ** 2

    # Create the plot
    plt.figure(figsize=(10, 6))
    plt.plot(x / 1000, y / 1000)  # Convert to km
    plt.title('Missile Trajectory')
    plt.xlabel('Distance (km)')
    plt.ylabel('Altitude (km)')
    plt.grid(True)
    plt.axhline(y=0, color='black', linestyle='-', alpha=0.3)

    # Save plot to a bytes buffer
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plt.close()

    return base64.b64encode(buffer.getvalue()).decode('utf-8')


def generate_velocity_time_graph(v_max, theta_deg):
    """Generate velocity-time graph"""
    g = 9.81
    theta_rad = np.radians(theta_deg)

    # Calculate flight time
    flight_time = (2 * v_max * np.sin(theta_rad)) / g

    # Create time points
    t = np.linspace(0, flight_time, 500)

    # Calculate velocity components
    v_x = v_max * np.cos(theta_rad) * np.ones_like(t)
    v_y = v_max * np.sin(theta_rad) - g * t
    v_total = np.sqrt(v_x ** 2 + v_y ** 2)

    # Create the plot
    plt.figure(figsize=(10, 6))
    plt.plot(t, v_x, label='Horizontal Velocity')
    plt.plot(t, v_y, label='Vertical Velocity')
    plt.plot(t, v_total, label='Total Velocity')
    plt.title('Velocity vs Time')
    plt.xlabel('Time (s)')
    plt.ylabel('Velocity (m/s)')
    plt.grid(True)
    plt.legend()

    # Save plot to a bytes buffer
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plt.close()

    return base64.b64encode(buffer.getvalue()).decode('utf-8')


def generate_stage_comparison_graph(delta_v_stages):
    """Generate stage comparison bar graph"""
    stages = [f"Stage {i + 1}" for i in range(len(delta_v_stages))]

    plt.figure(figsize=(8, 5))
    plt.bar(stages, delta_v_stages)
    plt.title('Delta-V Contribution by Stage')
    plt.xlabel('Rocket Stage')
    plt.ylabel('Delta-V (m/s)')
    plt.grid(True, axis='y')

    # Save plot to a bytes buffer
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plt.close()

    return base64.b64encode(buffer.getvalue()).decode('utf-8')


@app.route('/', methods=['GET', 'POST'])
def rocket_calculator():
    if request.method == 'POST':
        # Fixed parameters
        payload_kg = 250
        diameter_m = 0.75
        length_m = 7.5

        # Collect inputs from form
        propellant_density = float(request.form['propellant_density'])
        isp = float(request.form['isp'])

        # Collect booster lengths and propellant fractions
        booster_lengths = [
            float(request.form['booster1_length']),
            float(request.form['booster2_length']),
            float(request.form['booster3_length'])
        ]

        propellant_fractions = [
            float(request.form['booster1_fraction']),
            float(request.form['booster2_fraction']),
            float(request.form['booster3_fraction'])
        ]

        theta_deg = float(request.form['launch_angle'])
        # after you read theta_deg
        mode = request.form['mode']  # 'standard' or 'burntime'
        burn_time = float(request.form.get('burn_time', 0))

        # Check if total booster length matches missile length
        total_booster_length = sum(booster_lengths)
        if abs(total_booster_length - length_m) > 0.001:
            return render_template('index.html',
                                   error=f"Total booster length ({total_booster_length:.2f}m) does not match required missile length ({length_m}m)")

        # Calculate results
        results = calculate_rocket_parameters(
            payload_kg, diameter_m, length_m,
            propellant_density, isp,
            booster_lengths, propellant_fractions, theta_deg, mode=mode,
        burn_time=burn_time
        )

        # Generate graphs
        trajectory_img = generate_trajectory_graph(results['total_delta_v'], theta_deg)
        velocity_time_img = generate_velocity_time_graph(results['total_delta_v'], theta_deg)
        stage_comparison_img = generate_stage_comparison_graph(results['stage_delta_v'])

        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Rocket Calculation Results"

        # Add headers and results
        ws['A1'] = "Parameter"
        ws['B1'] = "Value"
        ws['A2'] = "Stage 1 Delta-V (m/s)"
        ws['B2'] = results['stage_delta_v'][0]
        ws['A3'] = "Stage 2 Delta-V (m/s)"
        ws['B3'] = results['stage_delta_v'][1]
        ws['A4'] = "Stage 3 Delta-V (m/s)"
        ws['B4'] = results['stage_delta_v'][2]
        ws['A5'] = "Total Delta-V (m/s)"
        ws['B5'] = results['total_delta_v']
        ws['A6'] = "Missile Range (km)"
        ws['B6'] = results['range_km']
        ws['A7'] = "Flight Time (s)"
        ws['B7'] = results['flight_time']

        # Add masses
        ws['A9'] = "Initial Masses"
        ws['A10'] = "Stage 1 Initial Mass (kg)"
        ws['B10'] = results['initial_masses'][0]
        ws['A11'] = "Stage 2 Initial Mass (kg)"
        ws['B11'] = results['initial_masses'][1]
        ws['A12'] = "Stage 3 Initial Mass (kg)"
        ws['B12'] = results['initial_masses'][2]

        ws['A14'] = "Propellant Masses"
        ws['A15'] = "Stage 1 Propellant Mass (kg)"
        ws['B15'] = results['propellant_masses'][0]
        ws['A16'] = "Stage 2 Propellant Mass (kg)"
        ws['B16'] = results['propellant_masses'][1]
        ws['A17'] = "Stage 3 Propellant Mass (kg)"
        ws['B17'] = results['propellant_masses'][2]

        # Save to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Display results and graphs in the template first
        return render_template('results.html',
                               results=results,
                               trajectory_img=trajectory_img,
                               velocity_time_img=velocity_time_img,
                               stage_comparison_img=stage_comparison_img)

    return render_template('index.html')


@app.route('/download-excel', methods=['POST'])
def download_excel():
    # Same calculation as above but directly return Excel file
    payload_kg = 250
    diameter_m = 0.75
    length_m = 7.5

    propellant_density = float(request.form['propellant_density'])
    isp = float(request.form['isp'])

    booster_lengths = [
        float(request.form['booster1_length']),
        float(request.form['booster2_length']),
        float(request.form['booster3_length'])
    ]

    propellant_fractions = [
        float(request.form['booster1_fraction']),
        float(request.form['booster2_fraction']),
        float(request.form['booster3_fraction'])
    ]

    theta_deg = float(request.form['launch_angle'])
    # after you read theta_deg
    mode = request.form['mode']  # 'standard' or 'burntime'
    burn_time = float(request.form.get('burn_time', 0))

    # Calculate results
    results = calculate_rocket_parameters(
        payload_kg, diameter_m, length_m,
        propellant_density, isp,
        booster_lengths, propellant_fractions, theta_deg, mode=mode,
        burn_time=burn_time
    )

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Rocket Calculation Results"

    # Add headers and results (same as before)
    ws['A1'] = "Parameter"
    ws['B1'] = "Value"
    ws['A2'] = "Stage 1 Delta-V (m/s)"
    ws['B2'] = results['stage_delta_v'][0]
    ws['A3'] = "Stage 2 Delta-V (m/s)"
    ws['B3'] = results['stage_delta_v'][1]
    ws['A4'] = "Stage 3 Delta-V (m/s)"
    ws['B4'] = results['stage_delta_v'][2]
    ws['A5'] = "Total Delta-V (m/s)"
    ws['B5'] = results['total_delta_v']
    ws['A6'] = "Missile Range (km)"
    ws['B6'] = results['range_km']
    ws['A7'] = "Flight Time (s)"
    ws['B7'] = results['flight_time']

    # Add masses
    ws['A9'] = "Initial Masses"
    ws['A10'] = "Stage 1 Initial Mass (kg)"
    ws['B10'] = results['initial_masses'][0]
    ws['A11'] = "Stage 2 Initial Mass (kg)"
    ws['B11'] = results['initial_masses'][1]
    ws['A12'] = "Stage 3 Initial Mass (kg)"
    ws['B12'] = results['initial_masses'][2]

    ws['A14'] = "Propellant Masses"
    ws['A15'] = "Stage 1 Propellant Mass (kg)"
    ws['B15'] = results['propellant_masses'][0]
    ws['A16'] = "Stage 2 Propellant Mass (kg)"
    ws['B16'] = results['propellant_masses'][1]
    ws['A17'] = "Stage 3 Propellant Mass (kg)"
    ws['B17'] = results['propellant_masses'][2]

    # Save to a bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='rocket_calculations.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def open_browser():
    webbrowser.open('http://127.0.0.1:5000')


if __name__ == '__main__':
    # Start browser after a short delay
    threading.Timer(1.5, open_browser).start()

    # Run the Flask app
    app.run(port=5000)
